# report_generator.py
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from datetime import datetime
import os
from database import Database

class ReportGenerator:
    def __init__(self):
        self.db = Database()
    
    def generate_daily_report(self, brand_id: str = None):
        """Generate comprehensive daily report"""
        
        # Get data
        orders = self._get_orders_data(brand_id)
        products = self._get_products_data(brand_id)
        sync_logs = self._get_sync_logs(brand_id)
        
        # Create Excel file
        filename = f"reports/jubelio_report_{datetime.now().strftime('%Y%m%d')}.xlsx"
        os.makedirs("reports", exist_ok=True)
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Sheet 1: Orders Summary
            orders_summary = orders.groupby(['brand_id', 'channel']).agg({
                'jubelio_order_id': 'count',
                'total_price': 'sum'
            }).rename(columns={'jubelio_order_id': 'order_count'})
            orders_summary.to_excel(writer, sheet_name='Orders Summary')
            
            # Sheet 2: Daily Orders
            orders['date'] = pd.to_datetime(orders['order_date']).dt.date
            daily_orders = orders.groupby(['brand_id', 'date']).agg({
                'jubelio_order_id': 'count',
                'total_price': 'sum'
            })
            daily_orders.to_excel(writer, sheet_name='Daily Orders')
            
            # Sheet 3: Products
            products.to_excel(writer, sheet_name='Products', index=False)
            
            # Sheet 4: Sync Logs
            sync_logs.to_excel(writer, sheet_name='Sync Logs', index=False)
            
            # Sheet 5: Top Products
            top_products = orders.groupby(['brand_id', 'sku']).size().reset_index(name='count')
            top_products = top_products.sort_values('count', ascending=False).head(20)
            top_products.to_excel(writer, sheet_name='Top Products', index=False)
        
        # Format Excel
        self._format_excel(filename)
        
        return filename
    
    def _get_orders_data(self, brand_id):
        query = """
            SELECT brand_id, jubelio_order_id, order_number, total_price, 
                   status, channel, customer_name, order_date
            FROM orders
            WHERE order_date >= CURRENT_DATE - INTERVAL '30 days'
        """
        if brand_id:
            query += f" AND brand_id = '{brand_id}'"
        
        cursor = self.db.execute_query(query)
        columns = ['brand_id', 'jubelio_order_id', 'order_number', 'total_price', 
                   'status', 'channel', 'customer_name', 'order_date']
        data = [dict(zip(columns, row)) for row in cursor.fetchall()]
        
        return pd.DataFrame(data)
    
    def _get_products_data(self, brand_id):
        query = "SELECT brand_id, sku, product_name, price, stock, category FROM products"
        if brand_id:
            query += f" WHERE brand_id = '{brand_id}'"
        
        cursor = self.db.execute_query(query)
        columns = ['brand_id', 'sku', 'product_name', 'price', 'stock', 'category']
        data = [dict(zip(columns, row)) for row in cursor.fetchall()]
        
        return pd.DataFrame(data)
    
    def _get_sync_logs(self, brand_id):
        query = """
            SELECT brand_id, data_type, status, records_count, 
                   error_message, started_at, completed_at
            FROM sync_logs
            WHERE started_at >= CURRENT_DATE - INTERVAL '7 days'
        """
        if brand_id:
            query += f" AND brand_id = '{brand_id}'"
        
        cursor = self.db.execute_query(query)
        columns = ['brand_id', 'data_type', 'status', 'records_count', 
                   'error_message', 'started_at', 'completed_at']
        data = [dict(zip(columns, row)) for row in cursor.fetchall()]
        
        return pd.DataFrame(data)
    
    def _format_excel(self, filename):
        """Apply formatting to Excel file"""
        wb = load_workbook(filename)
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Format headers
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filename)
    
    def generate_csv_report(self, report_type: str, brand_id: str = None):
        """Generate CSV report for specific data type"""
        
        if report_type == 'inventory':
            query = """
                SELECT brand_id, sku, product_name, price, stock, category
                FROM products
            """
            if brand_id:
                query += f" WHERE brand_id = '{brand_id}'"
            
            cursor = self.db.execute_query(query)
            columns = ['brand_id', 'sku', 'product_name', 'price', 'stock', 'category']
            data = [dict(zip(columns, row)) for row in cursor.fetchall()]
            df = pd.DataFrame(data)
            
        elif report_type == 'sales':
            query = """
                SELECT brand_id, DATE(order_date) as date, 
                       COUNT(*) as orders, SUM(total_price) as revenue
                FROM orders
                WHERE order_date >= CURRENT_DATE - INTERVAL '30 days'
            """
            if brand_id:
                query += f" AND brand_id = '{brand_id}'"
            query += " GROUP BY brand_id, DATE(order_date) ORDER BY date DESC"
            
            cursor = self.db.execute_query(query)
            columns = ['brand_id', 'date', 'orders', 'revenue']
            data = [dict(zip(columns, row)) for row in cursor.fetchall()]
            df = pd.DataFrame(data)
        
        else:
            raise ValueError(f"Unknown report type: {report_type}")
        
        filename = f"reports/{report_type}_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        os.makedirs("reports", exist_ok=True)
        df.to_csv(filename, index=False)
        
        return filename

# API endpoint for report generation
from fastapi import FastAPI, Query
from fastapi.responses import FileResponse

app = FastAPI()

@app.get("/api/report/daily")
async def generate_daily_report_api(
    brand_id: Optional[str] = Query(None),
    format: str = Query("excel", regex="^(excel|csv)$")
):
    generator = ReportGenerator()
    
    if format == "excel":
        filename = generator.generate_daily_report(brand_id)
        return FileResponse(
            filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=os.path.basename(filename)
        )
    else:
        filename = generator.generate_csv_report("sales", brand_id)
        return FileResponse(
            filename,
            media_type="text/csv",
            filename=os.path.basename(filename)
        )